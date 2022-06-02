# Imports -------------------
from tkinter import *  # For GUI
from random import choice  # For Computer to generate random numbers
from openpyxl import load_workbook, Workbook  # For adding score to excel sheet
from datetime import date, datetime  # To get date and time for the excel sheet


# Creating the window ----------
root = Tk()
root.title("Hand Cricket")
root.geometry("500x500")
root.config(bg="systemTransparent")


## Name Section ---------------
name_lbl = Label(root, text="Enter your name")
name_lbl.grid(row=0, column=0, columnspan=2)

name_ent = Entry(root)
name_ent.grid(row=0, column=2, columnspan=3)

# To keep the button disabled in the beginning
def copy_name():
    global name
    name = name_ent.get()
    for i in range(1, 7):
        globals()[f"opt_{i}"].config(state=NORMAL)  # Refer line


# This will get the name saved
tick_btn = Button(root, text="\u2705", command=copy_name)
tick_btn.grid(row=0, column=5)


# Computer's Score
comp_lbl = Label(root, text="Robot")
comp_lbl.grid(row=1, column=0)
comp_num = Label(root, text="")
comp_num.grid(row=1, column=1)


# Player's Score
ply_lbl = Label(root, text="Player")
ply_lbl.grid(row=2, column=0)
ply_num = Label(root, text="")
ply_num.grid(row=2, column=1)


score = 0  # This is the player's score

# To access number and its unicode value
num_opts = [
    [1, "\u0031\uFE0F\u20E3"],
    [2, "\u0032\uFE0F\u20E3"],
    [3, "\u0033\uFE0F\u20E3"],
    [4, "\u0034\uFE0F\u20E3"],
    [5, "\u0035\uFE0F\u20E3"],
    [6, "\u0036\uFE0F\u20E3"],
]

# This shows the final score
score_lbl = Label(root, text=f"")
score_lbl.grid(row=5, column=0, columnspan=5, pady=3)


# Each function for each options (1-6)
def fn_1():
    global score
    var = num_opts[0]
    ply_num.config(text=var[1])
    comp_num.config(text=choice(num_opts)[1])
    if score != 0:
        if comp_num.cget("text") == ply_num.cget("text"):
            score_lbl.config(text=f"your score is: {score}")
            score -= var[0]
            for i in range(1, 7):
                globals()[f"opt_{i}"].config(state=DISABLED)
    score += var[0]


def fn_2():
    global score
    var = num_opts[1]
    ply_num.config(text=var[1])
    comp_num.config(text=choice(num_opts)[1])
    if score != 0:
        if comp_num.cget("text") == ply_num.cget("text"):
            score_lbl.config(text=f"your score is: {score}")
            score -= var[0]
            for i in range(1, 7):
                globals()[f"opt_{i}"].config(state=DISABLED)
    score += var[0]


def fn_3():
    global score
    var = num_opts[2]
    ply_num.config(text=var[1])
    comp_num.config(text=choice(num_opts)[1])
    if score != 0:
        if comp_num.cget("text") == ply_num.cget("text"):
            score_lbl.config(text=f"your score is: {score}")
            score -= var[0]
            for i in range(1, 7):
                globals()[f"opt_{i}"].config(state=DISABLED)
    score += var[0]


def fn_4():
    global score
    var = num_opts[3]
    ply_num.config(text=var[1])
    comp_num.config(text=choice(num_opts)[1])
    if score != 0:
        if comp_num.cget("text") == ply_num.cget("text"):
            score_lbl.config(text=f"your score is: {score}")
            score -= var[0]
            for i in range(1, 7):
                globals()[f"opt_{i}"].config(state=DISABLED)
    score += var[0]


def fn_5():
    global score
    var = num_opts[4]
    ply_num.config(text=var[1])
    comp_num.config(text=choice(num_opts)[1])
    if score != 0:
        if comp_num.cget("text") == ply_num.cget("text"):
            score_lbl.config(text=f"your score is: {score}")
            score -= var[0]
            for i in range(1, 7):
                globals()[f"opt_{i}"].config(state=DISABLED)
    score += var[0]


def fn_6():
    global score
    var = num_opts[5]
    ply_num.config(text=var[1])
    comp_num.config(text=choice(num_opts)[1])
    if score != 0:
        if comp_num.cget("text") == ply_num.cget("text"):
            score_lbl.config(text=f"your score is: {score}")
            score -= var[0]
            for i in range(1, 7):
                globals()[f"opt_{i}"].config(state=DISABLED)
    score += var[0]


# Six buttons for six options
opt_1 = Button(root, text="\u0031\uFE0F\u20E3", command=fn_1, state=DISABLED)
opt_1.grid(row=3, column=0)
opt_2 = Button(root, text="\u0032\uFE0F\u20E3", command=fn_2, state=DISABLED)
opt_2.grid(row=3, column=1)
opt_3 = Button(root, text="\u0033\uFE0F\u20E3", command=fn_3, state=DISABLED)
opt_3.grid(row=3, column=2)
opt_4 = Button(root, text="\u0034\uFE0F\u20E3", command=fn_4, state=DISABLED)
opt_4.grid(row=3, column=3)
opt_5 = Button(root, text="\u0035\uFE0F\u20E3", command=fn_5, state=DISABLED)
opt_5.grid(row=3, column=4)
opt_6 = Button(root, text="\u0036\uFE0F\u20E3", command=fn_6, state=DISABLED)
opt_6.grid(row=3, column=5)


# Present date and time
now = datetime.now().strftime("%H:%M:%S")
today = date.today().strftime("%b-%d-%Y")  # Ex -> May-27-2022


root.mainloop()  # part of GUI


# Writing the scores in the spreadsheet
try:  # if the spreadsheet already exist (not the first time)
    wb = load_workbook("hand_cricket.xlsx")
    ws = wb.active


except FileNotFoundError:  # if the spreadsheet doesn't exist
    wb = Workbook()
    wb.save("hand_cricket.xlsx")
    ws = wb.active
    ws["A1"].value = "Name"
    ws["B1"].value = "Time"
    ws["C1"].value = "Date"
    ws["D1"].value = "Score"


if name != "":  # if name cell is not empty
    i = 2
    while i >= 2:  # starting from the second row
        if ws[f"A{i}"].value == None or ws[f"A{i}"].value == "":
            ws[f"A{i}"] = name
            ws[f"B{i}"] = now  # time
            ws[f"C{i}"] = today  # date
            ws[f"D{i}"] = score
            break
        i += 1  # If first row is not filled go to the second one
wb.save("hand_cricket.xlsx")  # Saving changes (important!!)
